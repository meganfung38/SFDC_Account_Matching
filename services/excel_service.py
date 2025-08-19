from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
import io
import pandas as pd

class ExcelService:
    """Service for Excel operations and Account data handling"""
    
    def __init__(self):
        # RingCentral Brand Colors
        self.rc_cerulean = "0684BC"      # RingCentral primary blue
        self.rc_orange = "FF7A00"        # RingCentral orange
        self.rc_ocean = "002855"         # RingCentral dark blue
        self.rc_linen = "F1EFEC"         # RingCentral background
        self.rc_ash = "C8C2B4"           # RingCentral light gray
        self.rc_warm_black = "2B2926"    # RingCentral dark gray
        
        # Excel Styling with RingCentral Colors
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color=self.rc_cerulean, end_color=self.rc_cerulean, fill_type="solid")
        self.summary_font = Font(bold=True, color=self.rc_ocean)
        self.summary_fill = PatternFill(start_color=self.rc_linen, end_color=self.rc_linen, fill_type="solid")
        self.title_font = Font(bold=True, size=16, color=self.rc_ocean)
        self.accent_fill = PatternFill(start_color=self.rc_orange, end_color=self.rc_orange, fill_type="solid")
        self.border = Border(
            left=Side(style='thin', color=self.rc_ash),
            right=Side(style='thin', color=self.rc_ash),
            top=Side(style='thin', color=self.rc_ash),
            bottom=Side(style='thin', color=self.rc_ash)
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.wrap_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    

    
    def parse_excel_file(self, file_content):
        """Parse uploaded Excel file and return sheet names and preview data"""
        try:
            # Load workbook from file content
            wb = load_workbook(io.BytesIO(file_content), read_only=True)
            sheet_names = wb.sheetnames
            
            # Get headers for ALL sheets (not just first one)
            all_headers = {}
            preview_data = []
            total_rows = 0
            
            for sheet_name in sheet_names:
                sheet = wb[sheet_name]
                
                # Get headers for this sheet
                sheet_headers = []
                first_row = next(sheet.iter_rows(values_only=True), None)
                if first_row:
                    sheet_headers = [cell if cell is not None else f"Column_{i+1}" for i, cell in enumerate(first_row)]
                
                all_headers[sheet_name] = sheet_headers
                
                # Only get preview data for the first sheet
                if sheet_name == sheet_names[0]:
                    for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                        if row_idx == 0:
                            continue  # Skip header row
                        elif row_idx < 11:  # First 10 data rows
                            row_data = [cell if cell is not None else "" for cell in row]
                            # Pad row to match header length
                            while len(row_data) < len(sheet_headers):
                                row_data.append("")
                            preview_data.append(row_data[:len(sheet_headers)])  # Trim to header length
                        else:
                            break
                    
                    # Calculate total rows for first sheet
                    total_rows = (sheet.max_row - 1) if sheet.max_row else 0
            
            wb.close()
            
            return {
                'success': True,
                'sheet_names': sheet_names,
                'headers': all_headers,  # Headers for ALL sheets
                'preview_data': preview_data,
                'total_rows': max(total_rows, len(preview_data))
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f"Error parsing Excel file: {str(e)}"
            }
    


    def extract_account_ids_from_excel(self, file_content, sheet_name, account_id_column):
        """Extract Account IDs from specified column in Excel file"""
        try:
            # Use pandas for easier data extraction - read as string to preserve Account ID format
            df = pd.read_excel(io.BytesIO(file_content), sheet_name=sheet_name, dtype={account_id_column: str})
            
            if account_id_column not in df.columns:
                return {
                    'success': False,
                    'error': f"Column '{account_id_column}' not found in sheet '{sheet_name}'"
                }
            
            # Extract Account IDs and remove null/empty values
            account_ids = df[account_id_column].dropna().astype(str).tolist()
            # Remove empty strings and whitespace-only strings, and handle Excel formatting issues
            cleaned_account_ids = []
            for aid in account_ids:
                aid_str = str(aid).strip()
                # Remove any Excel formatting artifacts
                if aid_str and aid_str.lower() not in ['nan', 'none', 'null']:
                    # Handle potential floating point conversion (e.g., "1.23456789012345e+17")
                    if 'e+' in aid_str.lower():
                        try:
                            # Convert scientific notation back to full number
                            aid_str = f"{float(aid_str):.0f}"
                        except:
                            pass
                    cleaned_account_ids.append(aid_str)
            
            account_ids = cleaned_account_ids
            
            # Get original data for later merging - handle NaN values
            # Replace NaN with empty string to avoid JSON serialization issues
            df_clean = df.where(pd.notnull(df), '')  # Replace NaN with empty string
            original_data = df_clean.to_dict('records')
            
            return {
                'success': True,
                'account_ids': account_ids,
                'original_data': original_data,
                'total_rows': len(df)
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f"Error extracting Account IDs: {str(e)}"
            }

    def create_basic_excel(self, data, headers, title="Data Export", filename_prefix="export"):
        """Create a basic Excel file with data and headers"""
        try:
            wb = Workbook()
            ws = wb.active
            if ws is not None:
                ws.title = "Data"
                
                # Add title
                current_row = 1
                ws.merge_cells(f'A{current_row}:{get_column_letter(len(headers))}{current_row}')
                ws[f'A{current_row}'] = title
                ws[f'A{current_row}'].font = self.title_font
                ws[f'A{current_row}'].alignment = self.center_alignment
                current_row += 1
                
                # Add timestamp
                ws.merge_cells(f'A{current_row}:{get_column_letter(len(headers))}{current_row}')
                ws[f'A{current_row}'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                ws[f'A{current_row}'].alignment = self.center_alignment
                current_row += 2
                
                # Add headers
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col, value=header)
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = self.center_alignment
                    cell.border = self.border
                
                current_row += 1
                
                # Add data rows
                for row_data in data:
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col, value=value)
                        cell.border = self.border
                    current_row += 1
                
                # Auto-adjust column widths
                for col in range(1, len(headers) + 1):
                    ws.column_dimensions[get_column_letter(col)].width = 20
            
            # Create file buffer
            file_buffer = io.BytesIO()
            wb.save(file_buffer)
            file_buffer.seek(0)
            
            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{filename_prefix}_{timestamp}.xlsx"
            
            return {
                'success': True,
                'file_buffer': file_buffer,
                'filename': filename
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f"Error creating Excel file: {str(e)}"
            } 

    def create_matching_results_export(self, matched_pairs, unmatched_customers=None, flagged_customers=None, invalid_customers=None, summary=None):
        """Create Excel export for dual-file matching results with one row per customer account"""
        try:
            wb = Workbook()
            
            # Sheet 1: Complete Customer Match Results (ALL customers)
            ws_results = wb.active
            ws_results.title = "Customer Match Results"
            
            # Add title
            current_row = 1
            title = "Customer-to-Shell Account Matching Results"
            ws_results.merge_cells(f'A{current_row}:S{current_row}')
            ws_results[f'A{current_row}'] = title
            ws_results[f'A{current_row}'].font = self.title_font
            ws_results[f'A{current_row}'].alignment = self.center_alignment
            ws_results[f'A{current_row}'].fill = PatternFill(start_color=self.rc_cerulean, end_color=self.rc_cerulean, fill_type="solid")
            current_row += 1
            
            # Add timestamp and summary
            ws_results.merge_cells(f'A{current_row}:S{current_row}')
            ws_results[f'A{current_row}'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws_results[f'A{current_row}'].alignment = self.center_alignment
            current_row += 1
            
            if summary:
                ws_results.merge_cells(f'A{current_row}:S{current_row}')
                summary_text = f"Total Customers: {summary.get('total_customer_accounts', 0)} | Matched: {summary.get('matched_pairs', 0)} | Unmatched: {summary.get('unmatched_customers', 0)} | Flagged: {summary.get('flagged_customer_accounts', 0)} | Processing Time: {summary.get('execution_time', 'N/A')}"
                ws_results[f'A{current_row}'] = summary_text
                ws_results[f'A{current_row}'].alignment = self.center_alignment
            current_row += 2
            
            # Headers for complete results table
            headers = [
                "Customer ID", "Customer Name", "Customer Website", "Customer Billing Address",
                "Match Status", "Match Reason",
                "Recommended Shell ID", "Shell Name", "Shell ZI ID", "Shell Website", "Shell Billing Address",
                "Overall Match Confidence", "Website Match Score", "Name Match Score", "Address Consistency Score",
                "AI Confidence Score", "AI Explanation",
                "Candidate Count", "Processing Notes"
            ]
            
            # Add headers
            for col, header in enumerate(headers, 1):
                cell = ws_results.cell(row=current_row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_alignment
                cell.border = self.border
            current_row += 1
            
            # Create a comprehensive list of ALL customers with their status
            all_customer_results = []
            
            # 1. Add matched customers
            for pair in matched_pairs:
                customer = pair['customer_account']
                shell = pair['recommended_shell']
                ai_assessment = pair.get('ai_assessment', {})
                
                all_customer_results.append({
                    'customer': customer,
                    'status': 'MATCHED',
                    'reason': f"Matched to shell account with {pair.get('match_confidence', 0):.1f}% confidence",
                    'shell': shell,
                    'match_confidence': pair.get('match_confidence', 0),
                    'website_match': pair.get('website_match', 0),
                    'name_match': pair.get('name_match', 0),
                    'address_consistency': pair.get('address_consistency', 0),
                    'ai_assessment': ai_assessment,
                    'candidate_count': pair.get('candidate_count', 0),
                    'total_shells': pair.get('total_shells', 0)
                })
            
            # 2. Add unmatched customers
            if unmatched_customers:
                for unmatched in unmatched_customers:
                    customer = unmatched['customer_account']
                    
                    all_customer_results.append({
                        'customer': customer,
                        'status': 'UNMATCHED',
                        'reason': unmatched.get('reason', 'No suitable shell match found'),
                        'shell': None,
                        'match_confidence': 0,
                        'website_match': 0,
                        'name_match': 0,
                        'address_consistency': 0,
                        'ai_assessment': {},
                        'candidate_count': 0,
                        'total_shells': 0
                    })
            
            # 3. Add flagged customers
            if flagged_customers:
                for flagged in flagged_customers:
                    bad_domain_info = flagged.get('Bad_Domain', {})
                    
                    all_customer_results.append({
                        'customer': flagged,
                        'status': 'FLAGGED',
                        'reason': f"Excluded from matching: {bad_domain_info.get('explanation', 'Bad domain detected')}",
                        'shell': None,
                        'match_confidence': 0,
                        'website_match': 0,
                        'name_match': 0,
                        'address_consistency': 0,
                        'ai_assessment': {},
                        'candidate_count': 0,
                        'total_shells': 0
                    })
            
            # 4. Add invalid customers
            if invalid_customers:
                for invalid_id in invalid_customers:
                    # Create a minimal customer object for invalid IDs
                    all_customer_results.append({
                        'customer': {'Id': invalid_id, 'Name': 'INVALID ACCOUNT ID', 'Website': '', 'BillingCity': '', 'BillingState': '', 'BillingCountry': '', 'BillingPostalCode': ''},
                        'status': 'INVALID',
                        'reason': 'Invalid Account ID - does not exist in Salesforce',
                        'shell': None,
                        'match_confidence': 0,
                        'website_match': 0,
                        'name_match': 0,
                        'address_consistency': 0,
                        'ai_assessment': {},
                        'candidate_count': 0,
                        'total_shells': 0
                    })
            
            # Sort results by customer name for better readability
            all_customer_results.sort(key=lambda x: x['customer'].get('Name', ''))
            
            # Add ALL customer results to the table
            for result in all_customer_results:
                customer = result['customer']
                shell = result['shell']
                ai_assessment = result['ai_assessment']
                
                # Format customer address
                customer_address_parts = []
                if customer.get('BillingCity'):
                    customer_address_parts.append(customer['BillingCity'])
                if customer.get('BillingState'):
                    customer_address_parts.append(customer['BillingState'])
                if customer.get('BillingCountry'):
                    customer_address_parts.append(customer['BillingCountry'])
                if customer.get('BillingPostalCode'):
                    customer_address_parts.append(customer['BillingPostalCode'])
                customer_address = ', '.join(customer_address_parts)
                
                # Format shell address (if matched)
                shell_address = ''
                if shell:
                    shell_address_parts = []
                    if shell.get('ZI_Company_City__c'):
                        shell_address_parts.append(shell['ZI_Company_City__c'])
                    if shell.get('ZI_Company_State__c'):
                        shell_address_parts.append(shell['ZI_Company_State__c'])
                    if shell.get('ZI_Company_Country__c'):
                        shell_address_parts.append(shell['ZI_Company_Country__c'])
                    if shell.get('ZI_Company_Postal_Code__c'):
                        shell_address_parts.append(shell['ZI_Company_Postal_Code__c'])
                    shell_address = ', '.join(shell_address_parts)
                
                # Format AI explanation
                ai_bullets = ai_assessment.get('explanation_bullets', [])
                ai_explanation = '\n'.join(ai_bullets) if ai_bullets else ('No AI analysis (not matched)' if result['status'] != 'MATCHED' else 'No AI explanation available')
                
                # Determine processing notes
                if result['status'] == 'MATCHED':
                    processing_notes = f"Evaluated {result['total_shells']} shell candidates, found {result['candidate_count']} potential matches"
                elif result['status'] == 'UNMATCHED':
                    processing_notes = "No matching shell candidates met minimum similarity threshold"
                else:  # FLAGGED
                    processing_notes = "Excluded from matching due to data quality issues"
                
                row_data = [
                    customer.get('Id', ''),
                    customer.get('Name', ''),
                    customer.get('Website', ''),
                    customer_address,
                    result['status'],
                    result['reason'],
                    shell.get('Id', '') if shell else '',
                    shell.get('ZI_Company_Name__c', '') if shell else '',
                    shell.get('ZI_Id__c', '') if shell else '',
                    shell.get('ZI_Website__c', '') if shell else '',
                    shell_address,
                    f"{result['match_confidence']:.1f}%" if result['match_confidence'] > 0 else '',
                    f"{result['website_match']:.1f}%" if result['website_match'] > 0 else '',
                    f"{result['name_match']:.1f}%" if result['name_match'] > 0 else '',
                    f"{result['address_consistency']:.1f}/100" if result['address_consistency'] > 0 else '',
                    f"{ai_assessment.get('confidence_score', 0)}/100" if ai_assessment.get('confidence_score', 0) > 0 else '',
                    ai_explanation,
                    result['candidate_count'] if result['candidate_count'] > 0 else '',
                    processing_notes
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws_results.cell(row=current_row, column=col, value=value)
                    cell.border = self.border
            
                    # Apply conditional formatting based on status
                    if result['status'] == 'MATCHED':
                        if col == 5:  # Status column
                            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
                        if col in [12, 13, 14, 15, 16]:  # Score columns
                            cell.alignment = self.center_alignment
                        else:
                            cell.alignment = self.wrap_alignment
                    elif result['status'] == 'UNMATCHED':
                        if col == 5:  # Status column
                            cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")  # Light orange
                        cell.alignment = self.wrap_alignment
                    else:  # FLAGGED
                        if col == 5:  # Status column
                            cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light red
                        cell.alignment = self.wrap_alignment
                
                current_row += 1
            
            # Auto-adjust column widths
            column_widths = {
                1: 18,   # Customer ID
                2: 30,   # Customer Name
                3: 25,   # Customer Website
                4: 35,   # Customer Address
                5: 12,   # Match Status
                6: 40,   # Match Reason
                7: 18,   # Shell ID
                8: 30,   # Shell Name
                9: 18,   # Shell ZI ID
                10: 25,  # Shell Website
                11: 35,  # Shell Address
                12: 15,  # Overall Confidence
                13: 15,  # Website Match
                14: 15,  # Name Match
                15: 15,  # Address Consistency
                16: 15,  # AI Confidence
                17: 50,  # AI Explanation
                18: 12,  # Candidate Count
                19: 40   # Processing Notes
            }
            
            for col, width in column_widths.items():
                column_letter = get_column_letter(col)
                ws_results.column_dimensions[column_letter].width = width
            
            # Freeze panes to keep headers visible
            ws_results.freeze_panes = "A2"
            
            # Sheet 2: Summary Metrics
            ws_summary = wb.create_sheet("Summary Metrics")
            
            # Add summary title
            current_row = 1
            ws_summary.merge_cells(f'A{current_row}:B{current_row}')
            ws_summary[f'A{current_row}'] = "Matching Process Summary"
            ws_summary[f'A{current_row}'].font = self.title_font
            ws_summary[f'A{current_row}'].alignment = self.center_alignment
            ws_summary[f'A{current_row}'].fill = PatternFill(start_color=self.rc_cerulean, end_color=self.rc_cerulean, fill_type="solid")
            current_row += 2
            
            # Summary metrics
            if summary:
                metrics = [
                    ["Total Customer Accounts Processed", summary.get('total_customer_accounts', 0)],
                    ["Successfully Matched", summary.get('matched_pairs', 0)],
                    ["Unable to Match", summary.get('unmatched_customers', 0)],
                    ["Flagged (Bad Domains)", summary.get('flagged_customer_accounts', 0)],
                    ["Total Shell Accounts Available", summary.get('total_shell_accounts', 0)],
                    ["Processing Time", summary.get('execution_time', 'N/A')],
                    ["Match Success Rate", f"{(summary.get('matched_pairs', 0) / max(summary.get('clean_customer_accounts', 1), 1) * 100):.1f}%" if summary.get('clean_customer_accounts', 0) > 0 else "0%"]
                ]
                
                # Add metric headers
                ws_summary[f'A{current_row}'] = "Metric"
                ws_summary[f'B{current_row}'] = "Value"
                ws_summary[f'A{current_row}'].font = self.header_font
                ws_summary[f'B{current_row}'].font = self.header_font
                ws_summary[f'A{current_row}'].fill = self.header_fill
                ws_summary[f'B{current_row}'].fill = self.header_fill
                current_row += 1
            
                # Add metrics
                for metric_name, metric_value in metrics:
                    ws_summary[f'A{current_row}'] = metric_name
                    ws_summary[f'B{current_row}'] = metric_value
                    ws_summary[f'A{current_row}'].border = self.border
                    ws_summary[f'B{current_row}'].border = self.border
            current_row += 1
            
            # Auto-adjust summary column widths
            ws_summary.column_dimensions['A'].width = 35
            ws_summary.column_dimensions['B'].width = 20
            
            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"customer_shell_matching_results_{timestamp}.xlsx"
            
            return {
                'success': True,
                'file_buffer': buffer,
                'filename': filename
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f"Error creating matching results export: {str(e)}"
            } 